import ijson
import requests
import datetime
import sys
import os
import json
import csv
import ipaddress
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from typing import Dict, Iterable, List, Optional, Tuple, Set, Any

LOG_FILE = "process_log.txt"
ASN_CACHE_FILE = "asn_cache.json"   # disk cache

PROBE_CSV = "probe_ids.csv"
PROBE_BATCH_SIZE = 20 

START_TIMESTAMP = 1692489600  # 2023-08-20
END_TIMESTAMP   = 1724112000  # 2024-08-20

# RIPE Atlas measurement IDs for root-servers (you can add others here)
ROOTSERVERS = [5001, 5004, 5005, 5006, 5008, 5009, 5010, 5011, 5012, 5013, 5014, 5015, 5016]
CAIDA_REL_FILE = "20240901.as-rel.txt"

FILTER_PROBE = 62292
HTTP_TIMEOUT = 30

# IXP prefix list file (your screenshot path)
IXP_PREFIXES_FILE = "/root/PROJECT/TRACE_ROUTE/trace_database/IXP/ixp-dataset/data/ixp_prefixes.txt"

asn_cache: Dict[str, Any] = {}
if os.path.exists(ASN_CACHE_FILE):
    with open(ASN_CACHE_FILE, "r") as f:
        try:
            asn_cache = json.load(f)
        except Exception:
            asn_cache = {}

def save_asn_cache():
    with open(ASN_CACHE_FILE, "w") as f:
        json.dump(asn_cache, f)

# -------------------
# Helpers
# -------------------
def split_time_range(start, end, delta_days=30):  # bigger chunks
    current = start
    while current < end:
        next_time = min(current + datetime.timedelta(days=delta_days), end)
        yield (current, next_time)
        current = next_time



def fetch_and_parse_json(url, probe_id):
    """
    Return list[(route_tuple, unix_ts)] for this probe/time window.
    We treat each route as the list of responding hops (strings of IPs).
    """
    try:
        resp = requests.get(
            url,
            stream=True,
            timeout=HTTP_TIMEOUT,
            headers={"Accept-Encoding": "gzip, deflate"}
        )
        if resp.status_code != 200:
            print(f"Failed to fetch data from {url}, Status code: {resp.status_code}")
            return []

        resp.raw.decode_content = True

        try:
            import ijson.backends.yajl2_c as ijson_backend
        except Exception:
            try:
                import ijson.backends.yajl2 as ijson_backend
            except Exception:
                import ijson.backends.python as ijson_backend

        out, seen = [], set()

        try:
            for obj in ijson_backend.items(resp.raw, 'item'):
                try:
                    ts = obj.get('timestamp')
                    if not ts:
                        continue
                    # Collect only responding hops; skip timeouts '*'
                    hops = []
                    for hop in obj.get('result', []):
                        if ('result' in hop and hop['result']
                                and isinstance(hop['result'], list)):
                            hop_from = hop['result'][0].get('from', '*')
                            if hop_from != '*':
                                hops.append(hop_from)
                    route = tuple(hops)
                    if route and (route, ts) not in seen:
                        seen.add((route, ts))
                        out.append((route, ts))
                except Exception as e:
                    print(f"Error processing object (probe {probe_id}): {e}")
            return out

        except Exception as e_stream:
            # Fallback to full JSON load for small windows
            try:
                data = resp.json()
            except Exception as e_json:
                print(f"Streaming parse failed ({e_stream}); fallback json() also failed: {e_json}")
                return []

            out, seen = [], set()
            for obj in data or []:
                try:
                    ts = obj.get('timestamp')
                    if not ts:
                        continue
                    hops = []
                    for hop in obj.get('result', []):
                        if ('result' in hop and hop['result']
                                and isinstance(hop['result'], list)):
                            hop_from = hop['result'][0].get('from', '*')
                            if hop_from != '*':
                                hops.append(hop_from)
                    route = tuple(hops)
                    if route and (route, ts) not in seen:
                        seen.add((route, ts))
                        out.append((route, ts))
                except Exception as e:
                    print(f"Error processing object (probe {probe_id}): {e}")
            return out

    except requests.RequestException as re:
        print(f"HTTP error fetching {url}: {re}")
        return []
    except Exception as e:
        print(f"Unexpected error fetching/parsing {url}: {e}")
        return []

def load_caida_relationships(filename):
    relationships = {}
    with open(filename, "r") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            parts = line.split("|")
            if len(parts) != 3:
                continue
            a1, a2, r = parts
           
            try:
                as1 = int(a1)
                as2 = int(a2)
                rel = int(r)
            except ValueError:
                continue

            if rel == -1:
                relationships[(as1, as2)] = "-1"
                relationships[(as2, as1)] = "-1"
            elif rel == 0:
                relationships[(as1, as2)] = "0"
                relationships[(as2, as1)] = "0"

    return relationships


def is_public_ip(ip: str) -> bool:
    """Return True if ip is a valid global (public) address; False for '*', RFC1918, link-local, etc."""
    if not ip or ip == '*':
        return False
    try:
        return ipaddress.ip_address(ip).is_global
    except ValueError:
        return False

def load_ixp_prefixes(file_path):
    with open(file_path, "r") as f:
        return [ipaddress.ip_network(line.strip(), strict=False) for line in f if line.strip()]

def is_ip_in_ixp(ip_str, ixp_networks):
    try:
        ip = ipaddress.ip_address(ip_str)
        for net in ixp_networks: # 'net' is already an ip_network object
            if ip in net: 
                return True
        return False
    except ValueError:
        return False

def get_asns(ip_add: str):
    """Query RIPEstat for ASN(s) covering the IP. Returns list or None."""
    # Simple cache
    if ip_add in asn_cache:
        return asn_cache[ip_add]
    url = f"https://stat.ripe.net/data/network-info/data.json?resource={ip_add}"
    try:
        response = requests.get(url, timeout=HTTP_TIMEOUT)
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and 'asns' in data['data']:
                asns = data['data']['asns']
                asn_cache[ip_add] = asns
                return asns
        else:
            print(f"Failed to fetch ASN data for IP: {ip_add} (HTTP {response.status_code})")
    except Exception as e:
        print(f"ASN lookup error for {ip_add}: {e}")
    return None

def get_single_asn(ip: str) -> Optional[int]:
    """Normalize get_asns(ip) to a single int ASN (take the first if multiple)."""
    asns = get_asns(ip)
    if not asns:
        return None
    try:
        return int(asns[0])
    except Exception:
        return None

# ---- Root recognition & relationship -----------------------------------------

def identify_root_server(dest_asn: Optional[int],
                         root_asn_map: Dict[str, Set[int]]) -> Optional[str]:
    """
    Map destination ASN to root name (a-root..m-root).
    root_asn_map example: {'k-root': {25152}, 'a-root': {19836}, ...}
    """
    if dest_asn is None:
        return None
    for root_name, asn_set in root_asn_map.items():
        if dest_asn in asn_set:
            return root_name
    return None


# ---- Main analysis ------------------------------------------------------------

def analyze_root_traceroutes(
    probe_ids: Iterable[int],
    start_timestamp: int,
    end_timestamp: int,
    base_url: str,
    caida_relationships,
    root_asn_map: Dict[str, Set[int]],
    ixp_prefixes
) -> List[Dict[str, Any]]:

    start_date = datetime.datetime.fromtimestamp(start_timestamp)
    end_date = datetime.datetime.fromtimestamp(end_timestamp)

    out: List[Dict[str, Any]] = []

    # Normalize probe_ids to iterable
    if isinstance(probe_ids, int):
        probe_ids = [probe_ids]

    for probe_id in probe_ids:
        for s_dt, e_dt in split_time_range(start_date, end_date):
            url = (
                f"{base_url}?probe_ids={probe_id}"
                f"&start={int(s_dt.timestamp())}"
                f"&stop={int(e_dt.timestamp())}"
                f"&format=json"
            )
            route_items: List[Tuple[List[str], int]] = fetch_and_parse_json(url, probe_id) or []

            for route, ts in route_items:
                # Need at least two responding hops for a penultimate
                if len(route) < 2:
                    continue

                dest_ip = route[-1]
                penult_ip = route[-2]

                # Penultimate must be a non-timeout public IP
                if not is_public_ip(penult_ip):
                    continue

                penult_asn = get_single_asn(penult_ip)
                if penult_asn is None:
                    continue

                dest_asn = get_single_asn(dest_ip)  # ASN of the root anycast hop
                root_name = identify_root_server(dest_asn, root_asn_map)

                if root_name is None:
                    # Not a recognized root-server dest (or ASN not in map)
                    continue

                # CAIDA relationship penultimate -> root
                relationship = "No Relationship"
                if dest_asn is not None:
                    relationship = caida_relationships.get((dest_asn, penult_asn), "No Relationship")
                    print("eheheh",relationship)

                # Exclude ASNs already peering (private or via IXP) with this root
                penult_in_ixp = is_ip_in_ixp(penult_ip, ixp_prefixes)
    
                out.append({
                    "probe_id": probe_id,
                    "root": root_name,
                    "dest_ip": dest_ip,
                    "dest_asn": dest_asn,
                    "penult_ip": penult_ip,
                    "penult_asn": penult_asn,
                    "relationship_penult_to_root": relationship,
                    "penult_in_ixp": penult_in_ixp,
                    "full_traceroute": route, 
                })
    return out

# -------------------
# Excel output
# -------------------
def _normalize_for_excel(v):
    if v is None:
        return ""
    if isinstance(v, (list, tuple, set)):
        return ", ".join(map(str, v))
    if isinstance(v, dict):
        return json.dumps(v, ensure_ascii=False)
    return v

def save_to_csv(results: List[Dict[str, Any]], filename: str):
    # same columns/order as the Excel export
    headers = [
        "probe_id",
        "root",
        "dest_ip",
        "dest_asn",
        "penult_ip",
        "penult_asn",
        "relationship_penult_to_root",
        "penult_in_ixp",
        "full_traceroute",
    ]

    # Convert list traceroutes to a single string (pipe-separated)
    def _fmt(v):
        if isinstance(v, (list, tuple)):
            return " | ".join(str(x) for x in v)
        return v

    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for r in results:
            row = {h: _fmt(r.get(h)) for h in headers}
            writer.writerow(row)

#XSLS SAVE PART --------------------------------------------
MAX_XLSX_ROWS = 1_048_576          # Excel hard limit
SAFE_ROWS_PER_FILE = 500_000       # chunk size to keep files responsive

def _write_sheet(ws, rows, headers):
    ws.title = "Penultimate Results"
    ws.append(headers)

    wrap = Alignment(wrap_text=True, vertical="top")

    # write rows; put each hop on its own line
    for r in rows:
        row_out = []
        for h in headers:
            if h == "full_traceroute":
                v = r.get(h, "")
                if isinstance(v, (list, tuple)):
                    v = "\n".join(str(x) for x in v)
                else:
                    v = str(v)
                row_out.append(v)
            else:
                row_out.append(_normalize_for_excel(r.get(h)))
        ws.append(row_out)

    # light formatting (no per-row height when many rows)
    widths = [12, 10, 15, 10, 15, 10, 18, 10, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # apply wrap; only compute row heights for small sheets (< 50k rows)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    if ws.max_row <= 50_000:
        full_col_idx = headers.index("full_traceroute") + 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ft_cell = row[full_col_idx - 1]
            lines = str(ft_cell.value).count("\n") + 1
            ws.row_dimensions[ft_cell.row].height = min(15 * lines, 600)

    ws.freeze_panes = "A2"


def save_to_xlsx_chunked(results: List[Dict[str, Any]], base_filename: str, rows_per_file: int = SAFE_ROWS_PER_FILE):
    headers = [
        "probe_id",
        "root",
        "dest_ip",
        "dest_asn",
        "penult_ip",
        "penult_asn",
        "relationship_penult_to_root",
        "penult_in_ixp",
        "full_traceroute",
    ]

    if not results:
        # still create an empty file for consistency
        wb = Workbook()
        _write_sheet(wb.active, [], headers)
        wb.save(base_filename)
        return [base_filename]

    files = []
    total = len(results)
    part = 1
    for start in range(0, total, rows_per_file):
        chunk = results[start:start + rows_per_file]
        wb = Workbook()
        _write_sheet(wb.active, chunk, headers)

        if start == 0:
            out = base_filename
        else:
            root, ext = os.path.splitext(base_filename)
            out = f"{root}_part{part}{ext}"
        wb.save(out)
        files.append(out)
        part += 1
    return files
#XSLS SAVE PART --------------------------------------------


def save_to_csv(results: List[Dict[str, Any]], filename: str):
    """
    Save results to CSV safely even for very large datasets.
    Each hop in full_traceroute will be joined by ' -> '.
    """
    headers = [
        "probe_id",
        "root",
        "dest_ip",
        "dest_asn",
        "penult_ip",
        "penult_asn",
        "relationship_penult_to_root",
        "penult_in_ixp",
        "full_traceroute",
    ]

    with open(filename, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for r in results:
            row = []
            for h in headers:
                val = r.get(h, "")
                if h == "full_traceroute":
                    if isinstance(val, (list, tuple)):
                        val = " -> ".join(str(x) for x in val)
                    else:
                        val = str(val)
                row.append(val)
            writer.writerow(row)


def load_root_asn_map() -> Dict[str, Set[int]]:
    """
    Fill this with authoritative ASN sets per root.
    Current entries are conservative; K-root is correct (AS25152).
    Add others as you verify (A-root=AS19836, C-root=AS19281, etc.).
    """
    return {
        "k-root": {25152},
        "a-root": {7342},
        "b-root": {394353},
        "c-root": {2149},
        "d-root": {10886},
        "e-root": {21556},
        "f-root": {3557},
        "g-root": {5927},
        "h-root": {1508},
        "i-root": {29216},
        "j-root": {26415},
        "l-root": {20144},
        "m-root": {7500},  
    }

def load_probe_ids_from_csv(csv_file, batch_number, batch_size=PROBE_BATCH_SIZE):
    probe_ids = []
    with open(csv_file, mode='r') as file:
        reader = csv.reader(file)
        next(reader)  # Skip header
        start_index = (batch_number - 1) * batch_size
        for idx, row in enumerate(reader):
            if idx >= start_index and idx < start_index + batch_size:
                probe_ids.append(int(row[0]))  # Assuming probe_id is in the first column
    return probe_ids

# -------------------
# Main
# -------------------
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 trace_batch.py <output_folder> <batch_number>")
        sys.exit(1)

    folder = sys.argv[1]
    batch_number = int(sys.argv[2])
    os.makedirs(folder, exist_ok=True)

    # Load CAIDA rels (directed)
    caida_relationships = load_caida_relationships(CAIDA_REL_FILE)

    # Load IXP prefixes
    ixp_prefixes = load_ixp_prefixes(IXP_PREFIXES_FILE)
    print(f"[INFO] Loaded {len(ixp_prefixes)} IXP prefixes")

    # Root → ASN map & peering filters
    root_asn_map = load_root_asn_map()
    probe_ids = load_probe_ids_from_csv(PROBE_CSV, batch_number, PROBE_BATCH_SIZE)
    if not probe_ids:
        print(f"[WARN] No probe IDs found for batch {batch_number}")
        sys.exit(0)
    print(f"[INFO] Batch {batch_number}: {len(probe_ids)} probes → {probe_ids[:5]}{'...' if len(probe_ids)>5 else ''}")

    all_results: List[Dict[str, Any]] = []
    for measurement_id in ROOTSERVERS:
        base_url = f"https://atlas.ripe.net/api/v2/measurements/{measurement_id}/results/"
        print(f"[INFO] Processing measurement {measurement_id} for {len(probe_ids)} probes")
        res = analyze_root_traceroutes(
            probe_ids,
            START_TIMESTAMP,
            END_TIMESTAMP,
            base_url,
            caida_relationships,
            root_asn_map,
            ixp_prefixes
        )
        all_results.extend(res)

    out_file = os.path.join(folder, f"penultimate_results_batch_{batch_number}.xlsx")
    out_file_csv = os.path.join(folder, f"penultimate_results_batch_{batch_number}.csv")

    save_to_xlsx_chunked(all_results, out_file)  # (chunked XLSX version)
    save_to_csv(all_results, out_file_csv)
    save_asn_cache()
    print(f"[DONE] Saved {len(all_results)} rows to {out_file}")
